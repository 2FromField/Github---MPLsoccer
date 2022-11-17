import glob
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill

def getStatsPerSeasonsL1(globPath):
    """
    Generated an Excel file containing the stats of each UberEat Premier League team per season from 2009-10 to 2018-19. 
    
    Parameters:
        globPath (str) : path to the folder containing all the csv files of the regular seasons
        
    Returns: 
        Excel file with a sheet for each season coloured for each aggregate (best statistics in red and worst in blue)
    """
    
    Seasons = []
    df_AllSeasons = pd.DataFrame([],columns=[])
    #
    with pd.ExcelWriter('Ligue1_Stats.xlsx', engine='xlsxwriter') as writer:
        for file in glob.glob(f'{globPath}*.csv'):
            df = pd.read_csv(file)
            print(file)
            
            # List of teams per Season
            teams = df.HomeTeam.tolist() + df.AwayTeam.tolist()
            teams = sorted(list(set(teams)))
            
                
            df_perSeason = pd.DataFrame([],columns=[])
            line = 0
            #
            for k in range(len(teams)):
                
                df_perSeason.loc[line,'Team'] = teams[k]
                df_perSeason.loc[line,'Season'] = f"20{df.Date[0][-2:]}-{df.Date[len(df)-1][-2:]}"

                # Filtered dataframe
                df_all = df[(df.HomeTeam == teams[k])| (df.AwayTeam == teams[k])].reset_index(drop=True)  # GENERAL STATS
                df_home = df[(df.HomeTeam == teams[k])].reset_index(drop=True)                            # HOME STATS
                df_away = df[(df.AwayTeam == teams[k])].reset_index(drop=True)                            # AWAY STATS
                

                ## Win Rate
                homeWin = 0
                homeDefeat = 0
                homeDraw = 0
                awayWin = 0
                awayDefeat = 0
                awayDraw = 0
                #
                for h in range(len(df_home)):
                    if df_home['FTHG'][h] > df_home['FTAG'][h]:
                        homeWin += 1
                    elif df_home['FTHG'][h] == df_home['FTAG'][h]:
                        homeDraw += 1
                    else:
                        homeDefeat += 1
                #
                for h in range(len(df_away)):
                    if df_away['FTAG'][h] > df_away['FTHG'][h]:
                        awayWin += 1
                    elif df_away['FTAG'][h] == df_away['FTHG'][h]:
                        awayDraw += 1
                    else:
                        awayDefeat += 1
                #
                df_perSeason.loc[line,'Total_win'] = homeWin + awayWin
                df_perSeason.loc[line,'Total_defeat'] = homeDefeat + awayDefeat
                df_perSeason.loc[line,'Total_draw'] = homeDraw + awayDraw
                df_perSeason.loc[line,'Home_win'] = homeWin
                df_perSeason.loc[line,'Home_defeat'] = homeDefeat
                df_perSeason.loc[line,'Away_win'] = awayWin
                df_perSeason.loc[line,'Away_defeat'] = awayDefeat
                df_perSeason.loc[line,'Draw'] = homeDraw + awayDraw

                ## GOAL
                df_perSeason.loc[line,'Away_goal'] = df_away['FTAG'].sum()
                df_perSeason.loc[line,'Away_goal_perMatch'] = round((df_away['FTAG'].sum())/len(df_away),1)
                df_perSeason.loc[line,'Home_goal'] = df_home['FTHG'].sum()
                df_perSeason.loc[line,'Home_goal_perMatch'] = round((df_home['FTHG'].sum())/len(df_home),1)
                df_perSeason.loc[line,'Total_goal'] = df_home['FTHG'].sum() + df_away['FTAG'].sum()
                df_perSeason.loc[line,'Total_goal_perMatch'] = round((df_home['FTHG'].sum() + df_away['FTAG'].sum())/(len(df_home) + len(df_away)),1)
                df_perSeason.loc[line,'Half_goal'] = df_home['HTHG'].sum() + df_away['HTAG'].sum() 
                df_perSeason.loc[line,'Half_goal_perMatch'] = round((df_home['HTHG'].sum() + df_away['HTAG'].sum())/(len(df_home) + len(df_away)),1)

                ## Shots
                ##### HOME
                df_perSeason.loc[line,'Home_shots'] = df_home['HS'].sum()
                df_perSeason.loc[line,'ShotOnGoal_ratio'] = round((df_home['HST'].sum()*100)/df_home['HS'].sum(),1)
                df_perSeason.loc[line,'Home_ShotsOnTarget'] = df_home['HST'].sum()
                df_perSeason.loc[line,'Home_DecisifShotsOnTarget'] = round((df_home['FTHG'].sum()*100)/df_home['HST'].sum(),1)
                ##### AWAY
                df_perSeason.loc[line,'Away_shots'] = df_away['AS'].sum()
                df_perSeason.loc[line,'ShotOnGoal_ratio'] = round((df_away['AST'].sum()*100)/df_away['AS'].sum(),1)
                df_perSeason.loc[line,'Away_ShotsOnTarget'] = df_away['AST'].sum()
                df_perSeason.loc[line,'Away_DecisifShotsOnTarget'] = round((df_away['FTAG'].sum()*100)/df_away['AST'].sum(),1)

                ## Corners
                df_perSeason.loc[line,'Home_corners'] = df_home['HC'].sum()
                df_perSeason.loc[line,'Home_corners_perMatch'] = round(df_home['HC'].sum()/len(df_home),1)
                df_perSeason.loc[line,'Away_corners'] = df_away['AC'].sum()
                df_perSeason.loc[line,'Away_corners_perMatch'] = round(df_away['AC'].sum()/len(df_away),1)

                ## Fouls (R/Y)
                df_perSeason.loc[line,'Home_fouls'] = df_home['HF'].sum()
                df_perSeason.loc[line,'Home_fouls_perMatch'] = round(df_away['AF'].sum()/len(df_away),1)
                df_perSeason.loc[line,'Yellow_cards'] = df_away['AY'].sum() + df_home['HY'].sum()
                df_perSeason.loc[line,'Red_cards'] = df_away['AR'].sum() + df_home['HR'].sum()

                df_AllSeasons.loc[line,f"{df_perSeason.Season[0]}"] = df_perSeason.Total_win[line]
                
                line += 1
            
            # Add each season to the list of seasons travelled
            Seasons.append(df_perSeason.Season[0])
            
            # Sorted by win rate
            df_perSeason = df_perSeason.sort_values(by=['Total_win'],ascending=False).reset_index(drop=True)  

            # Sauvegarde du fichier R2
            df_perSeason.to_excel(writer,sheet_name=f"{df_perSeason.Season[0]}", startcol=-1)
            

    # Browse all the Season's sheet and add the colour
    for s in Seasons:     
        # Load the excel file
        wb = openpyxl.load_workbook('Ligue1_Stats.xlsx')
        ws = wb[f"{s}"]
        df = pd.read_excel('Ligue1_Stats.xlsx', sheet_name=f"{s}")

        # Font Color
        max = PatternFill(patternType='solid', fgColor='ff5252')
        min = PatternFill(patternType='solid', fgColor='62d7ff')

        # Alphabet from 'A' to 'ZZ'
        alphabet = [chr(h) for h in range(ord('A'),ord('Z')+1)]
        for h in np.arange(0,26,1):
            alphabet.append(f"{alphabet[0]}{alphabet[h]}")
        
        # MINS & MAXS lists -> color per lines
        Maxs = []
        Mins = []
        for i in range(len(df.columns)-2):
            Maxs.append(df.query(f"{df.columns[i+2]} == {df[str(df.columns[i+2])].max()}").index.tolist())
            Mins.append(df.query(f"{df.columns[i+2]} == {df[str(df.columns[i+2])].min()}").index.tolist())

        for col in range(len(df.columns)-2):
            df.rename(columns={f"{df.columns[col]}":f"{alphabet[col]}"})
            for k in range(len(Maxs[col])):
                ws[f"{alphabet[col+2]}{Maxs[col][k]+2}"].fill = max
            for h in range(len(Mins[col])):
                ws[f"{alphabet[col+2]}{Mins[col][h]+2}"].fill = min

        # Saved the excel files and the colors
        wb.save('Ligue1_Stats.xlsx')


getStatsPerSeasonsL1('Season_csv/')